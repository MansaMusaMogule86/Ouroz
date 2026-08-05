import json
import re
from collections import defaultdict
from pathlib import Path
import openpyxl

SOURCE_DOCUMENT = 'danat-al-jazeera-catalog'
SUPPLIER = {
    'name': 'Danat Al Jazeera',
    'slug': 'danat-al-jazeera',
    'verification_status': 'verified_draft',
    'description': 'B2B and wholesale supplier draft catalog imported from supplier document extract.'
}

SHEETS = ['GROCERIES', 'KITCHEN_ACCESSORIES', 'Accessories']

brand_tokens = [
    'dalfour','jibal','itkane','itkan','perly','merendina','safi','fino','alsa','amlou'
]

pack_re = re.compile(r'(\d+\s*(?:kg|g|gm|ml|l|pcs|pc|cm)(?:\s*[xX]\s*\d+)?)|(\d+\s*[xX]\s*\d+\s*(?:kg|g|gm|ml|l|pcs|pc))', re.IGNORECASE)
carton_re = re.compile(r'\b(?:x|X)\s*(\d+)\b')


def slugify(text: str) -> str:
    text = (text or '').lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    text = re.sub(r'-+', '-', text).strip('-')
    return text or 'unnamed-product'


def normalize_base_name(name: str) -> str:
    if not name:
        return ''
    s = name.lower().strip()
    s = re.sub(r'\([^)]*\)', ' ', s)
    s = re.sub(r'\d+\s*(kg|g|gm|ml|l|pcs|pc|cm)', ' ', s)
    s = re.sub(r'\b\d+\s*[xX]\s*\d+\b', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip(' -')
    return s


def detect_brand(text: str):
    t = (text or '').lower()
    for token in brand_tokens:
        if token in t:
            return token.title().replace('Itkane', 'Al Itkane').replace('Itkan', 'Al Itkane')
    return None


def classify_group(name: str, sheet: str):
    t = (name or '').lower()

    if any(k in t for k in ['tea', 'coffee', 'nescafe', 'cafe']):
        return 'Tea and coffee'
    if 'sugar' in t or '???' in t:
        return 'Sugar'
    if any(k in t for k in ['tuna', 'sardine']):
        return 'Tuna and sardines'
    if any(k in t for k in ['biscuit', 'chocolate', 'cookie', 'merendina']):
        return 'Biscuits and chocolate'
    if any(k in t for k in ['couscous', 'pasta', 'semolina', 'smid', 'flour', 'fino']):
        return 'Couscous, pasta, semolina, and flour'
    if any(k in t for k in ['syrup', 'beverage', 'drink', 'vinegar', 'jam', 'tomato paste']):
        return 'Syrups, beverages, vinegar, jam, and tomato paste'
    if any(k in t for k in ['olive', 'preserved lemon', 'harissa', 'khleaa', 'pickled']):
        return 'Olives, preserved lemon, harissa, and khleaa'
    if any(k in t for k in ['baking', 'dessert', 'yeast', 'cake', 'custard']):
        return 'Baking and dessert products'
    if any(k in t for k in ['sauce', 'dressing', 'ketchup', 'mayo', 'mustard']):
        return 'Sauces and dressings'
    if any(k in t for k in ['tagine', 'ceramic', 'plate', 'mug']):
        return 'Tagines and ceramics'
    if any(k in t for k in ['tea cup', 'tea tray', 'couscous pot', 'teapot', 'tea pot', 'couscous']):
        return 'Tea cups, tea trays, couscous pots, and teapots'

    if sheet in ['KITCHEN_ACCESSORIES', 'Accessories']:
        return 'Tagines and ceramics'

    return None


def category_from_group(group: str):
    if group in ['Tagines and ceramics', 'Tea cups, tea trays, couscous pots, and teapots']:
        return 'kitchen-accessories'
    return 'groceries'


wb = openpyxl.load_workbook('OUROZ.xlsx', data_only=True)
entries = []

for sheet in SHEETS:
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        original = (row[0] if len(row) > 0 else None) or ''
        english = (row[1] if len(row) > 1 else None) or ''
        arabic = (row[2] if len(row) > 2 else None) or ''

        if not original and not english and not arabic:
            continue

        canonical_name = (english or original).strip()
        if not canonical_name:
            continue

        group = classify_group(canonical_name, sheet)
        if not group:
            # still import as unclear for manual review
            group = 'UNMAPPED_MANUAL_REVIEW'

        pack_matches = [m.group(0) for m in pack_re.finditer(canonical_name)]
        pack_size = pack_matches[0].strip() if pack_matches else None

        carton_match = carton_re.search(canonical_name)
        carton_qty = carton_match.group(1) if carton_match else None

        ambiguous = False
        reasons = []
        if not english and original:
            reasons.append('missing_english_name')
        if not arabic:
            reasons.append('missing_arabic_name')
        if group == 'UNMAPPED_MANUAL_REVIEW':
            reasons.append('unmapped_group')
        if len(canonical_name) < 4 or 'product' in canonical_name.lower():
            ambiguous = True
            reasons.append('ambiguous_or_generic_name')

        entries.append({
            'sheet': sheet,
            'row': idx,
            'original_name': original.strip() or None,
            'english_name': english.strip() or None,
            'arabic_name': arabic.strip() or None,
            'canonical_name': canonical_name,
            'group': group,
            'category_slug': category_from_group(group) if group != 'UNMAPPED_MANUAL_REVIEW' else 'manual-review',
            'brand': detect_brand(canonical_name),
            'pack_size': pack_size,
            'carton_quantity': carton_qty,
            'price_status': 'pending_supplier_confirmation',
            'retail_availability_status': 'pending_supplier_confirmation',
            'wholesale_availability_status': 'pending_supplier_confirmation',
            'source_document': SOURCE_DOCUMENT,
            'source_page': None,
            'source_reference': f'{sheet}:row-{idx}',
            'manual_review_required': ambiguous or bool(reasons),
            'manual_review_reasons': reasons,
        })

# aggregate to products + variants
products_map = {}
for e in entries:
    key_base = normalize_base_name(e['canonical_name']) or normalize_base_name(e['original_name'] or '') or e['canonical_name'].lower()
    key = (key_base, e['group'], e['category_slug'])

    if key not in products_map:
        pname = e['english_name'] or e['original_name'] or e['canonical_name']
        products_map[key] = {
            'id': f"daj-{slugify(key_base)}",
            'slug': slugify(key_base),
            'english_name': pname,
            'arabic_name': e['arabic_name'],
            'original_name': e['original_name'],
            'brand': e['brand'],
            'product_group': e['group'],
            'product_category': e['category_slug'],
            'supplier_slug': SUPPLIER['slug'],
            'price_status': 'pending_supplier_confirmation',
            'retail_availability_status': 'pending_supplier_confirmation',
            'wholesale_availability_status': 'pending_supplier_confirmation',
            'source_document': SOURCE_DOCUMENT,
            'source_page': None,
            'source_references': [],
            'manual_review_required': False,
            'manual_review_reasons': [],
            'variants': []
        }

    v_slug = slugify((e['pack_size'] or 'unspecified') + '-' + e['source_reference'])
    products_map[key]['variants'].append({
        'id': f"{products_map[key]['id']}-var-{v_slug}",
        'pack_size': e['pack_size'],
        'carton_quantity': e['carton_quantity'],
        'moq': None,
        'price_status': 'pending_supplier_confirmation',
        'source_document': SOURCE_DOCUMENT,
        'source_page': e['source_page'],
        'source_reference': e['source_reference'],
    })

    products_map[key]['source_references'].append(e['source_reference'])
    if e['manual_review_required']:
        products_map[key]['manual_review_required'] = True
        for r in e['manual_review_reasons']:
            if r not in products_map[key]['manual_review_reasons']:
                products_map[key]['manual_review_reasons'].append(r)

# dedupe variants by pack_size/carton combination
for p in products_map.values():
    seen = set()
    deduped = []
    for v in p['variants']:
        k = (v['pack_size'], v['carton_quantity'])
        if k in seen:
            continue
        seen.add(k)
        deduped.append(v)
    p['variants'] = deduped

products = sorted(products_map.values(), key=lambda x: (x['product_group'], x['english_name']))

quality = {
    'source_document': SOURCE_DOCUMENT,
    'supplier': SUPPLIER['name'],
    'total_rows_imported': len(entries),
    'total_products': len(products),
    'total_variants': sum(len(p['variants']) for p in products),
    'manual_review_products': sum(1 for p in products if p['manual_review_required']),
    'unmapped_group_products': sum(1 for p in products if p['product_group'] == 'UNMAPPED_MANUAL_REVIEW'),
    'missing_arabic_name_products': sum(1 for p in products if 'missing_arabic_name' in p['manual_review_reasons']),
    'missing_english_name_products': sum(1 for p in products if 'missing_english_name' in p['manual_review_reasons']),
    'missing_page_reference_products': len(products),
}

payload = {
    'supplier': SUPPLIER,
    'product_groups': [
        'Tea and coffee',
        'Sugar',
        'Tuna and sardines',
        'Biscuits and chocolate',
        'Couscous, pasta, semolina, and flour',
        'Syrups, beverages, vinegar, jam, and tomato paste',
        'Olives, preserved lemon, harissa, and khleaa',
        'Baking and dessert products',
        'Sauces and dressings',
        'Tagines and ceramics',
        'Tea cups, tea trays, couscous pots, and teapots',
        'UNMAPPED_MANUAL_REVIEW'
    ],
    'quality': quality,
    'products': products,
}

out = Path('src/lib/catalog/danatAlJazeeraCatalog.generated.json')
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'WROTE {out} with {len(products)} products and {quality["total_variants"]} variants')
